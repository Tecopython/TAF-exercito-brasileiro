from openpyxl import load_workbook
from pathlib import Path


# tabelas masculino LEMB
dicio_corrida_m_b = {
    "18-21": {
        "I": (0, 2599),
        "R": (2600, 2799),
        "B": (2800, 3149),
        "MB": (3150, 3199),
        "E": (3200, float('inf'))
    },
    "22-25": {
        "I": (0, 2699),
        "R": (2700, 2849),
        "B": (2850, 3099),
        "MB": (3100, 3249),
        "E": (3250, float('inf'))
    },
    "26-29": {
        "I": (0, 2599),
        "R": (2600, 2749),
        "B": (2750, 2999),
        "MB": (3000, 3149),
        "E": (3150, float('inf'))
    },
    "30-33": {
        "I": (0, 2549),
        "R": (2550, 2649),
        "B": (2650, 2899),
        "MB": (2900, 3099),
        "E": (3100, float('inf'))
    },
    "34-37": {
        "I": (0, 2449),
        "R": (2450, 2549),
        "B": (2550, 2799),
        "MB": (2800, 2949),
        "E": (2950, float('inf'))
    },
    "38-41": {
        "I": (0, 2349),
        "R": (2350, 2449),
        "B": (2450, 2699),
        "MB": (2700, 2849),
        "E": (2850, float('inf'))
    },
    "42-45": {
        "I": (0, 2249),
        "R": (2250, 2399),
        "B": (2400, 2599),
        "MB": (2600, 2749),
        "E": (2750, float('inf'))
    },
    "46-49": {
        "I": (0, 2149),
        "R": (2150, 2299),
        "B": (2300, 2499),
        "MB": (2500, 2649),
        "E": (2650, float('inf'))
    },
    "50-53": 1900,
    "54-57": 1800,
    "58-61": 1600,
    "62-65": 1400
}

dicio_flexao_m_b = {
    "18-21": {
        "I": (0, 21),
        "R": (22, 24),
        "B": (25, 33),
        "MB": (34, 38),
        "E": (39, float('inf'))
    },
    "22-25": {
        "I": (0, 23),
        "R": (24, 26),
        "B": (27, 35),
        "MB": (36, 40),
        "E": (41, float('inf'))
    },
    "26-29": {
        "I": (0, 21),
        "R": (22, 24),
        "B": (25, 33),
        "MB": (34, 38),
        "E": (39, float('inf'))
    },
    "30-33": {
        "I": (0, 20),
        "R": (21, 23),
        "B": (24, 31),
        "MB": (32, 36),
        "E": (37, float('inf'))
    },
    "34-37": {
        "I": (0, 17),
        "R": (18, 20),
        "B": (21, 28),
        "MB": (29, 33),
        "E": (34, float('inf'))
    },
    "38-41": {
        "I": (0, 16),
        "R": (17, 19),
        "B": (20, 27),
        "MB": (28, 31),
        "E": (32, float('inf'))
    },
    "42-45": {
        "I": (0, 14),
        "R": (15, 17),
        "B": (18, 24),
        "MB": (25, 28),
        "E": (29, float('inf'))
    },
    "46-49": {
        "I": (0, 11),
        "R": (12, 14),
        "B": (15, 21),
        "MB": (22, 25),
        "E": (26, float('inf'))
    },
    "50-53": 11,
    "54-57": 9,
    "58-61": 8,
    "62-65": 6
}

dicio_abdominal_m_b = {
    "18-21": {
        "I": (0, 34),
        "R": (35, 44),
        "B": (45, 63),
        "MB": (64, 73),
        "E": (74, float('inf'))
    },
    "22-25": {
        "I": (0, 41),
        "R": (42, 51),
        "B": (52, 68),
        "MB": (69, 78),
        "E": (79, float('inf'))
    },
    "26-29": {
        "I": (0, 37),
        "R": (38, 48),
        "B": (49, 65),
        "MB": (66, 75),
        "E": (76, float('inf'))
    },
    "30-33": {
        "I": (0, 33),
        "R": (34, 42),
        "B": (43, 60),
        "MB": (61, 69),
        "E": (70, float('inf'))
    },
    "34-37": {
        "I": (0, 30),
        "R": (31, 39),
        "B": (40, 56),
        "MB": (57, 65),
        "E": (66, float('inf'))
    },
    "38-41": {
        "I": (0, 28),
        "R": (29, 37),
        "B": (38, 54),
        "MB": (55, 63),
        "E": (64, float('inf'))
    },
    "42-45": {
        "I": (0, 26),
        "R": (27, 35),
        "B": (36, 52),
        "MB": (53, 61),
        "E": (62, float('inf'))
    },
    "46-49": {
        "I": (0, 24),
        "R": (25, 33),
        "B": (34, 50),
        "MB": (51, 59),
        "E": (60, float('inf'))
    },
    "50-53": 23,
    "54-57": 21,
    "58-61": 19,
    "62-65": 17
}

dicio_barra_m_b = {
    "18-21": {
        "I": (0, 4),
        "R": (5, 6),
        "B": (7, 9),
        "MB": (10, 11),
        "E": (12, float('inf'))
    },
    "22-25": {
        "I": (0, 5),
        "R": (6, 7),
        "B": (8, 10),
        "MB": (11, 12),
        "E": (13, float('inf'))
    },
    "26-29": {
        "I": (0, 4),
        "R": (5, 6),
        "B": (7, 9),
        "MB": (10, 11),
        "E": (12, float('inf'))
    },
    "30-33": {
        "I": (0, 4),
        "R": (5, 5),
        "B": (6, 8),
        "MB": (9, 10),
        "E": (11, float('inf'))
    },
    "34-37": {
        "I": (0, 3),
        "R": (4, 4),
        "B": (5, 6),
        "MB": (7, 8),
        "E": (9, float('inf'))
    },
    "38-39": {
        "I": (0, 2),
        "R": (3, 3),
        "B": (4, 5),
        "MB": (6, 7),
        "E": (8, float('inf'))
    },
    "40-45":2,
    "46-49":1,
    "50-70":0
}

#  tabelas femininas LEMB
dicio_corrida_f_b = {
    "18-21": {
        "I": (0, 2099),
        "R": (2100, 2199),
        "B": (2200, 2449),
        "MB": (2450, 2599),
        "E": (2600, float('inf'))
    },
    "22-25": {
        "I": (0, 2149),
        "R": (2150, 2249),
        "B": (2250, 2449),
        "MB": (2450, 2649),
        "E": (2650, float('inf'))
    },
    "26-29": {
        "I": (0, 2099),
        "R": (2100, 2199),
        "B": (2200, 2399),
        "MB": (2400, 2599),
        "E": (2600, float('inf'))
    },
    "30-33": {
        "I": (0, 2049),
        "R": (2050, 2149),
        "B": (2150, 2349),
        "MB": (2350, 2549),
        "E": (2550, float('inf'))
    },
    "34-37": {
        "I": (0, 1999),
        "R": (2000, 2099),
        "B": (2100, 2299),
        "MB": (2300, 2499),
        "E": (2500, float('inf'))
    },
    "38-41": {
        "I": (0, 1899),
        "R": (1900, 1999),
        "B": (2000, 2199),
        "MB": (2200, 2399),
        "E": (2400, float('inf'))
    },
    "42-45": {
        "I": (0, 1849),
        "R": (1850, 1949),
        "B": (1950, 2149),
        "MB": (2150, 2299),
        "E": (2300, float('inf'))
    },
    "46-49": {
        "I": (0, 1749),
        "R": (1750, 1849),
        "B": (1850, 2049),
        "MB": (2050, 2249),
        "E": (2250, float('inf'))
    },
    "50-53": 1600,
    "54-57": 1500,
    "58-61": 1300,
    "62-65": 1100
}

dicio_flexao_f_b = {
    "18-21": {
        "I": (0, 10),
        "R": (11, 11),
        "B": (12, 16),
        "MB": (17, 19),
        "E": (20, float('inf'))
    },
    "22-25": {
        "I": (0, 11),
        "R": (12, 12),
        "B": (13, 18),
        "MB": (19, 21),
        "E": (22, float('inf'))
    },
    "26-29": {
        "I": (0, 10),
        "R": (11, 11),
        "B": (12, 16),
        "MB": (17, 19),
        "E": (20, float('inf'))
    },
    "30-33": {
        "I": (0, 9),
        "R": (10, 10),
        "B": (11, 15),
        "MB": (16, 18),
        "E": (19, float('inf'))
    },
    "34-37": {
        "I": (0, 8),
        "R": (9, 9),
        "B": (10, 14),
        "MB": (15, 17),
        "E": (18, float('inf'))
    },
    "38-41": {
        "I": (0, 7),
        "R": (8, 8),
        "B": (9, 13),
        "MB": (14, 16),
        "E": (17, float('inf'))
    },
    "42-45": {
        "I": (0, 6),
        "R": (7, 7),
        "B": (8, 12),
        "MB": (13, 15),
        "E": (16, float('inf'))
    },
    "46-49": {
        "I": (0, 5),
        "R": (6, 6),
        "B": (7, 10),
        "MB": (11, 13),
        "E": (14, float('inf'))
    },
    "50-53": 5,
    "54-57": 4,
    "58-61": 3,
    "62-65": 2
}

dicio_abdominal_f_b = {
    "18-21": {
        "I": (0, 30),
        "R": (31, 39),
        "B": (40, 55),
        "MB": (56, 64),
        "E": (65, float('inf'))
    },
    "22-25": {
        "I": (0, 32),
        "R": (33, 41),
        "B": (42, 57),
        "MB": (58, 66),
        "E": (67, float('inf'))
    },
    "26-29": {
        "I": (0, 31),
        "R": (32, 40),
        "B": (41, 56),
        "MB": (57, 65),
        "E": (66, float('inf'))
    },
    "30-33": {
        "I": (0, 29),
        "R": (30, 38),
        "B": (39, 54),
        "MB": (55, 63),
        "E": (64, float('inf'))
    },
    "34-37": {
        "I": (0, 27),
        "R": (28, 36),
        "B": (37, 52),
        "MB": (53, 61),
        "E": (62, float('inf'))
    },
    "38-41": {
        "I": (0, 25),
        "R": (26, 34),
        "B": (35, 50),
        "MB": (51, 59),
        "E": (60, float('inf'))
    },
    "42-45": {
        "I": (0, 23),
        "R": (24, 32),
        "B": (33, 48),
        "MB": (49, 57),
        "E": (58, float('inf'))
    },
    "46-49": {
        "I": (0, 21),
        "R": (22, 30),
        "B": (31, 46),
        "MB": (47, 55),
        "E": (56, float('inf'))
    },
    "50-53": 20,
    "54-57": 18,
    "58-61": 16,
    "62-65": 14
}

dicio_barra_f_b = {
    "18-21": {
        "I": (0, 0),
        "R": (1, 2),
        "B": (3, 4),
        "MB": (5, 5),
        "E": (6, float('inf'))
    },
    "22-25": {
        "I": (0, 1),
        "R": (2, 3),
        "B": (4, 5),
        "MB": (6, 6),
        "E": (7, float('inf'))
    },
    "26-29": {
        "I": (0, 1),
        "R": (2, 2),
        "B": (3, 4),
        "MB": (5, 5),
        "E": (6, float('inf'))
    },
    "30-33": {
        "I": (0, 0),
        "R": (1, 2),
        "B": (3, 4),
        "MB": (5, 5),
        "E": (6, float('inf'))
    },
    "34-37": {
        "I": (0, 0),
        "R": (1, 2),
        "B": (3, 3),
        "MB": (4, 4),
        "E": (5, float('inf'))
    },
    "38-39": {
        "I": (0, 0),
        "R": (1, 1),
        "B": (2, 2),
        "MB": (3, 3),
        "E": (4, float('inf'))
    },
    "40-45":45,
    "46-49":30
}

# tabela masculina complementar
dicio_corrida_m_c = {
    "18-21": {
        "I": (0, 2599),
        "R": (2600, 2699),
        "B": (2700, 2899),
        "MB": (2900, 2999),
        "E": (3000, float('inf'))
    },
    "22-25": {
        "I": (0, 2699),
        "R": (2700, 2799),
        "B": (2800, 2949),
        "MB": (2950, 3049),
        "E": (3050, float('inf'))
    },
    "26-29": {
        "I": (0, 2599),
        "R": (2600, 2699),
        "B": (2700, 2849),
        "MB": (2850, 2949),
        "E": (2950, float('inf'))
    },
    "30-33": {
        "I": (0, 2549),
        "R": (2550, 2649),
        "B": (2650, 2799),
        "MB": (2800, 2899),
        "E": (2900, float('inf'))
    },
    "34-37": {
        "I": (0, 2449),
        "R": (2450, 2549),
        "B": (2550, 2649),
        "MB": (2650, 2749),
        "E": (2750, float('inf'))
    },
    "38-41": {
        "I": (0, 2349),
        "R": (2350, 2449),
        "B": (2450, 2549),
        "MB": (2550, 2649),
        "E": (2650, float('inf'))
    },
    "42-45": {
        "I": (0, 2249),
        "R": (2250, 2349),
        "B": (2350, 2499),
        "MB": (2500, 2599),
        "E": (2600, float('inf'))
    },
    "46-49": {
        "I": (0, 2149),
        "R": (2150, 2299),
        "B": (2300, 2399),
        "MB": (2400, 2499),
        "E": (2500, float('inf'))
    },
    "50-53": 1900,
    "54-57": 1800,
    "58-61": 1600,
    "62-65": 1400
}

dicio_flexao_m_c = {
    "18-21": {
        "I": (0, 21),
        "R": (22, 24),
        "B": (25, 29),
        "MB": (30, 32),
        "E": (33, float('inf'))
    },
    "22-25": {
        "I": (0, 23),
        "R": (24, 26),
        "B": (27, 31),
        "MB": (32, 34),
        "E": (35, float('inf'))
    },
    "26-29": {
        "I": (0, 21),
        "R": (22, 24),
        "B": (25, 29),
        "MB": (30, 32),
        "E": (33, float('inf'))
    },
    "30-33": {
        "I": (0, 20),
        "R": (21, 23),
        "B": (24, 27),
        "MB": (28, 30),
        "E": (31, float('inf'))
    },
    "34-37": {
        "I": (0, 17),
        "R": (18, 20),
        "B": (21, 25),
        "MB": (26, 28),
        "E": (29, float('inf'))
    },
    "38-41": {
        "I": (0, 16),
        "R": (17, 19),
        "B": (20, 23),
        "MB": (24, 26),
        "E": (27, float('inf'))
    },
    "42-45": {
        "I": (0, 14),
        "R": (15, 17),
        "B": (18, 21),
        "MB": (22, 24),
        "E": (25, float('inf'))
    },
    "46-49": {
        "I": (0, 11),
        "R": (12, 14),
        "B": (15, 18),
        "MB": (19, 21),
        "E": (22, float('inf'))
    },
    "50-53": 11,
    "54-57": 9,
    "58-61": 8,
    "62-65": 6
}

dicio_abdominal_m_c = {
    "18-21": {
        "I": (0, 34),
        "R": (35, 44),
        "B": (45, 63),
        "MB": (64, 73),
        "E": (74, float('inf'))
    },
    "22-25": {
        "I": (0, 41),
        "R": (42, 51),
        "B": (52, 68),
        "MB": (69, 78),
        "E": (79, float('inf'))
    },
    "26-29": {
        "I": (0, 37),
        "R": (38, 48),
        "B": (49, 65),
        "MB": (66, 75),
        "E": (76, float('inf'))
    },
    "30-33": {
        "I": (0, 33),
        "R": (34, 42),
        "B": (43, 60),
        "MB": (61, 69),
        "E": (70, float('inf'))
    },
    "34-37": {
        "I": (0, 30),
        "R": (31, 39),
        "B": (40, 56),
        "MB": (57, 65),
        "E": (66, float('inf'))
    },
    "38-41": {
        "I": (0, 28),
        "R": (29, 37),
        "B": (38, 54),
        "MB": (55, 63),
        "E": (64, float('inf'))
    },
    "42-45": {
        "I": (0, 26),
        "R": (27, 35),
        "B": (36, 52),
        "MB": (53, 61),
        "E": (62, float('inf'))
    },
    "46-49": {
        "I": (0, 24),
        "R": (25, 33),
        "B": (34, 50),
        "MB": (51, 59),
        "E": (60, float('inf'))
    },
    "50-53": 23,
    "54-57": 21,
    "58-61": 19,
    "62-65": 17
}

#Tabela feminica complementar
dicio_corrida_f_c = {
    "18-21": {
        "I": (0, 2099),
        "R": (2100, 2199),
        "B": (2200, 2299),
        "MB": (2300, 2399),
        "E": (2400, float('inf'))
    },
    "22-25": {
        "I": (0, 2149),
        "R": (2150, 2249),
        "B": (2250, 2349),
        "MB": (2350, 2449),
        "E": (2450, float('inf'))
    },
    "26-29": {
        "I": (0, 2099),
        "R": (2100, 2199),
        "B": (2200, 2299),
        "MB": (2300, 2399),
        "E": (2400, float('inf'))
    },
    "30-33": {
        "I": (0, 2049),
        "R": (2050, 2149),
        "B": (2150, 2249),
        "MB": (2250, 2349),
        "E": (2350, float('inf'))
    },
    "34-37": {
        "I": (0, 1999),
        "R": (2000, 2099),
        "B": (2100, 2199),
        "MB": (2200, 2299),
        "E": (2300, float('inf'))
    },
    "38-41": {
        "I": (0, 1899),
        "R": (1900, 1999),
        "B": (2000, 2149),
        "MB": (2150, 2249),
        "E": (2250, float('inf'))
    },
    "42-45": {
        "I": (0, 1849),
        "R": (1850, 1949),
        "B": (1950, 2049),
        "MB": (2050, 2149),
        "E": (2150, float('inf'))
    },
    "46-49": {
        "I": (0, 1749),
        "R": (1750, 1849),
        "B": (1850, 1949),
        "MB": (1950, 2049),
        "E": (2050, float('inf'))
    },
    "50-53": 1600,
    "54-57": 1500,
    "58-61": 1300,
    "62-65": 1100
}

dicio_flexao_f_c = {
    "18-21": {
        "I": (0, 10),
        "R": (11, 11),
        "B": (12, 14),
        "MB": (15, 15),
        "E": (16, float('inf'))
    },
    "22-25": {
        "I": (0, 11),
        "R": (12, 12),
        "B": (13, 15),
        "MB": (16, 16),
        "E": (17, float('inf'))
    },
    "26-29": {
        "I": (0, 10),
        "R": (11, 11),
        "B": (12, 14),
        "MB": (15, 15),
        "E": (16, float('inf'))
    },
    "30-33": {
        "I": (0, 9),
        "R": (10, 10),
        "B": (11, 13),
        "MB": (14, 14),
        "E": (15, float('inf'))
    },
    "34-37": {
        "I": (0, 8),
        "R": (9, 9),
        "B": (10, 12),
        "MB": (13, 13),
        "E": (14, float('inf'))
    },
    "38-41": {
        "I": (0, 7),
        "R": (8, 8),
        "B": (9, 11),
        "MB": (12, 12),
        "E": (13, float('inf'))
    },
    "42-45": {
        "I": (0, 6),
        "R": (7, 7),
        "B": (8, 10),
        "MB": (11, 11),
        "E": (12, float('inf'))
    },
    "46-49": {
        "I": (0, 5),
        "R": (6, 6),
        "B": (7, 9),
        "MB": (10, 10),
        "E": (11, float('inf'))
    },
    "50-53": 5,
    "54-57": 4,
    "58-61": 3,
    "62-65": 2
}

dicio_abdominal_f_c = {
    "18-21": {
        "I": (0, 30),
        "R": (31, 39),
        "B": (40, 55),
        "MB": (56, 64),
        "E": (65, float('inf'))
    },
    "22-25": {
        "I": (0, 32),
        "R": (33, 41),
        "B": (42, 57),
        "MB": (58, 66),
        "E": (67, float('inf'))
    },
    "26-29": {
        "I": (0, 31),
        "R": (32, 40),
        "B": (41, 56),
        "MB": (57, 65),
        "E": (66, float('inf'))
    },
    "30-33": {
        "I": (0, 29),
        "R": (30, 38),
        "B": (39, 54),
        "MB": (55, 63),
        "E": (64, float('inf'))
    },
    "34-37": {
        "I": (0, 27),
        "R": (28, 36),
        "B": (37, 52),
        "MB": (53, 61),
        "E": (62, float('inf'))
    },
    "38-41": {
        "I": (0, 25),
        "R": (26, 34),
        "B": (35, 50),
        "MB": (51, 59),
        "E": (60, float('inf'))
    },
    "42-45": {
        "I": (0, 23),
        "R": (24, 32),
        "B": (33, 48),
        "MB": (49, 57),
        "E": (58, float('inf'))
    },
    "46-49": {
        "I": (0, 21),
        "R": (22, 30),
        "B": (31, 46),
        "MB": (47, 55),
        "E": (56, float('inf'))
    },
    "50-53": 19,
    "54-57": 17,
    "58-61": 15,
    "62-65": 13
}


# Dicionário que mapeia o nome da atividade para as tabelas correspondentes (masculino e feminino)
dicio_atividades = {'B':{
    "CORRIDA": {
        "M": dicio_corrida_m_b,
        "F": dicio_corrida_f_b
    },
    "FLEXAO": {
        "M": dicio_flexao_m_b,
        "F": dicio_flexao_f_b
    },
    "ABDOMINAL": {
        "M": dicio_abdominal_m_b,
        "F": dicio_abdominal_f_b
    },
    "BARRA": {
        "M": dicio_barra_m_b,
        "F": dicio_barra_f_b
    }
},
'CT':{
    "CORRIDA": {
        "M": dicio_corrida_m_c,
        "F": dicio_corrida_f_c
    },
    "FLEXAO": {
        "M": dicio_flexao_m_c,
        "F": dicio_flexao_f_c
    },
    "ABDOMINAL": {
        "M": dicio_abdominal_m_c,
        "F": dicio_abdominal_f_c
    }
}
}

def determinar_mencao(idade,dicio_atividades,atividade,lem,segmento,indice):# retorna a menção correta para o índice indicado
    '''segmento - retirado da coluna 'SEGMENTO' da tabela
       idade - retirado da coluna "IDADE" da tabela
       atividade - nome da coluna da atividade na tabela
       desempenho - item retirado da coluna da atividade
       lem - retirado da coluna "LEM"
    '''
    if idade == None or isinstance(idade,str) or idade < 18:
        return 'erro idade'

    if segmento not in ['M','F'] or segmento == None:
        return 'erro no segmento'

    if lem not in ['B','CT'] or lem == None:
        return 'erro na LEM'

    if indice == None:
        return f'erro no indice - {atividade}'

    if isinstance(indice,str):
        if indice == 'NR' or indice == 'A':
            return indice
        else:
            return f'erro no indice - {atividade}'
    indice = int(indice)

# determinar a faixa de idade
    if lem == 'B':
            for faixa in dicio_atividades[lem][atividade][segmento].keys():
                inicio, fim = map(int, faixa.split('-'))
                if inicio <= idade <= fim:
                    faixa = faixa
                    break
    else:
        if atividade == 'BARRA':
            faixa = 'X'
        else:
            for faixa in dicio_atividades[lem][atividade][segmento].keys():
                inicio, fim = map(int, faixa.split('-'))
                if inicio <= idade <= fim:
                    faixa = faixa
                    break

#determinar a menção
    if faixa != 'X':
        for f in dicio_atividades[lem][atividade][segmento].keys():
            if f == faixa:
                dicio_mencoes = dicio_atividades[lem][atividade][segmento][f]
                if isinstance(dicio_mencoes, dict):
                    for k, (min,max) in dicio_mencoes.items():
                        if min <= indice <= max:
                            return k
                else:
                    if indice >= dicio_mencoes:
                        return "S"
                    else:
                        return "I"
    else:
        return indice

#Testar a menção
#determinar_mencao(50,dicio_atividades,'BARRA','CT','M',0)
# print(type(tabela['LEM'][29]))
# print(tabela['NOME'][4])


def lista_mencoes(aba):
    dicionario = dict()
    for linha in range(2, aba.max_row+1):
        lista = list()
        # pega os índices
        idade = aba.cell(row=linha, column=3).value
        segmento = aba.cell(row=linha, column=4).value
        lem = aba.cell(row=linha, column=5).value
        corrida = aba.cell(row=linha, column=6).value
        flexao = aba.cell(row=linha, column=7).value
        abdominal = aba.cell(row=linha, column=8).value
        barra = aba.cell(row=linha, column=9).value
        #pega a menção
        lista.append(determinar_mencao(idade,dicio_atividades,'CORRIDA',lem,segmento,corrida))
        lista.append(determinar_mencao(idade,dicio_atividades,'FLEXAO',lem,segmento,flexao))
        lista.append(determinar_mencao(idade,dicio_atividades,'ABDOMINAL',lem,segmento,abdominal))
        lista.append(determinar_mencao(idade,dicio_atividades,'BARRA',lem,segmento,barra))
        dicionario[f'{aba.cell(row=linha, column=1).value}'f' {aba.cell(row=linha, column=2).value}'] = lista
    return dicionario

#Teste da função
#print(lista_mencoes(aba))

def verifica_erros_lancamento(dicio):#Inserir a função lista_mencoes() dentro # retorna um dicionario com os erros.
    dicionario = dict() 
    #lista_verificacao = ['I','R','B','MB','E','S','NR','X','A']
    for k,v in dicio.items():
        if 'erro idade' in v:
            dicionario[k] = 'erro idade'
            continue
        if 'erro no segmento' in v:
            dicionario[k] = 'erro no segmento'
            continue
        if 'erro na LEM' in v:
            dicionario[k] = 'erro na LEM'
            continue
        if 'erro no indice - CORRIDA' in v:
            dicionario[k] = 'erro no lançamento da CORRIDA'
            continue
        if 'erro no indice da FLEXAO' in v:
            dicionario[k] = 'erro no lançamento da FLEXÃO'
            continue
        if 'erro no indice - ABDOMINAL ' in v:
            dicionario[k] = 'erro no lançamento do ABDOMINAL'
            continue
        if 'erro no indice - BARRA' in v:
            dicionario[k] = 'erro no lançamento da BARRA'
            continue
    return dicionario

#teste da função
#print(verifica_erros_lancamento(lista_mencoes(aba)))

def limpa_mencao(dicio,dicio_erros):#Inserir a função mencoes_por_item() dentro + a função verifica_erros()
    for k in dicio_erros.keys():
        dicio.pop(k)
    return dicio


#teste função
#print(limpa_mencao(lista_mencoes(aba),verifica_erros_lancamento(lista_mencoes(aba))))



def mencao_final(dicio):
    resultado = dict()
    for k, v in dicio.items():
        if 'erro no segmento' in v:
            resultado[k] = 'erro no segmento'
            continue
        if 'erro na LEM' in v:
            resultado[k] = 'erro na LEM'
            continue
        if 'erro idade' in v:
            resultado[k] = 'erro idade'
            continue
        if 'erro no indice - CORRIDA' in v:
            resultado[k] = 'erro no lançamento da CORRIDA'
            continue
        if 'erro no indice - FLEXAO' in v:
            resultado[k] = 'erro no lançamento da FLEXÃO'
            continue
        if 'erro no indice - ABDOMINAL' in v:
            resultado[k] = 'erro no lançamento do ABDOMINAL'
            continue
        if 'erro no indice - BARRA' in v:
            resultado[k] = 'erro no lançamento da BARRA'
            continue
        if 'NR' in v:
            resultado[k] = 'NR'
            continue
        if 'A' in v:
            if 'I' in v:
                resultado[k] = 'I'
                continue
            else:
                resultado[k] = 'R'
                continue
        if 'I' in v:
            resultado[k] = 'I'
            continue
        if 'R' in v:
            resultado[k] = 'R'
            continue
        if 'B' in v:
            resultado[k] = 'B'
            continue
        if 'MB' in v:
            resultado[k] = 'MB'
            continue
        if 'E' in v:
            resultado[k] = 'E'
            continue
        if 'S' in v:
            resultado[k] = 'S'
    return resultado

#teste da função
#mencao_final_correta = mencao_final(lista_mencoes(aba))

def mencao_final_limpa(dicio): #menção final sem os militares com erros de lançamento
    resultado = dict()         #mencao_final_limpa(lista_mencoes(aba))
    for k, v in dicio.items():
        if 'NR' in v:
            resultado[k] = 'NR'
            continue
        if 'A' in v:
            if 'I' in v:
                resultado[k] = 'I'
                continue
            else:
                resultado[k] = 'R'
                continue
        if 'I' in v:
            resultado[k] = 'I'
            continue
        if 'R' in v:
            resultado[k] = 'R'
            continue
        if 'B' in v:
            resultado[k] = 'B'
            continue
        if 'MB' in v:
            resultado[k] = 'MB'
            continue
        if 'E' in v:
            resultado[k] = 'E'
            continue
        if 'S' in v:
            resultado[k] = 'S'
    return resultado

def mencao_final_erros(dicio):
    resultado = dict()
    for k, v in dicio.items():
        if 'erro no segmento' in v:
            resultado[k] = 'erro no segmento'
            continue
        if 'erro na LEM' in v:
            resultado[k] = 'erro na LEM'
            continue
        if 'erro idade' in v:
            resultado[k] = 'erro idade'
            continue
        if 'erro no indice - CORRIDA' in v:
            resultado[k] = 'erro no lançamento da CORRIDA'
            continue
        if 'erro no indice - FLEXAO' in v:
            resultado[k] = 'erro no lançamento da FLEXÃO'
            continue
        if 'erro no indice - ABDOMINAL' in v:
            resultado[k] = 'erro no lançamento do ABDOMINAL'
            continue
        if 'erro no indice - BARRA' in v:
            resultado[k] = 'erro no lançamento da BARRA'
            continue
    return resultado


def mencao_lancada(aba):
    dicionario = dict()
    for linha in range(2, aba.max_row+1):
        pg_nome = f'{aba.cell(row=linha, column=1).value} {aba.cell(row=linha, column=2).value}'
        mencao = aba.cell(row=linha, column=10).value
        dicionario[pg_nome] = mencao
    return dicionario
#pg_nome = f'{aba.cell(row=2, column=1).value} {aba.cell(row=2, column=2).value}'
#mencao_lancada_tabela = mencao_lancada(tabela)
#print(mencao_lancada(aba))


def erros_lancamento(mencao_l, mencao_apurada):
    lista = list()
    for k,v in mencao_l.items():
        try:
            if mencao_apurada[k] != v:
                print(f'{k} está com a menção {v}, mas a correta é {mencao_apurada[k]}')
                lista.append(v)
            else:
                print('',end='')
        except:
            #print(f'{k} não pode ser verificado pois possui algum erro de lançamento')
            print('',end='')
    if lista == []:
        print('*'*70)
        print('''     NÃO FORAM ENCONTRADOS ERROS NAS MENÇÕES FINAIS
DOS MILITARES QUE ESTAVAM SEM ALTERAÇÃO NO LANÇAMENTO.''')
        print('*'*70)
#para teste da função
#erros_lancamento(mencao_lancada(aba),mencao_final(lista_mencoes(aba)))




def pega_aba(nome_arquivo):
    cont = 0
    while True:
        print('''Digite no número do TAF para consulta:
        1 - para o 1º TAF
        2 - para o 2º TAF
        3 - para o 3º TAF  ''')
        try:
            escolha = int(input('Escolha ->  '))
            if escolha == 1:
                sheet = '1º TAF'
            elif escolha == 2:
                sheet = '2º TAF'
            elif escolha == 3:
                sheet = '3º TAF'
            else:
                print('Digite um número interiro válido, entre 1 e 3: ')
                input('Aperte Enter para continuar')
                continue
            #PEGANDO O ARQUIVO
            planilha = load_workbook(nome_arquivo)
            aba = planilha[sheet]
            #TROCANDO O NOME DAS COLUNAS
            aba['G1'].value = 'FLEXAO'
            aba['J1'].value = 'MENCAO'
            linha_remover = list()
            for linha in aba.iter_rows():
                if all(celula.value is None or celula.value == "" for celula in linha):
                    linha_remover.append(linha[0].row)
            for linha_num in reversed(linha_remover):
                aba.delete_rows(linha_num)
            #REMOÇÃO DE COLUNAS VAZIAS
            coluna_remover = list()
            for coluna in aba.iter_cols():
                if all(celula.value is None or celula.value == "" for celula in coluna):
                    coluna_remover.append(coluna[0].column)
            for coluna_nun in reversed(coluna_remover):
                aba.delete_cols(coluna_nun)
            return aba
        except:
            if cont < 1:
                print('\nDigite um número interiro válido, entre 1 e 3: ')
                cont += 1
                continue
            else:
                print('''Caso o erro persista, verifique na planilha:
    1) Se o nome do arquivo está todo em maiúsculo - "PLANILHA TAF.xlsx" - esse é o nome correto.
    2) Se a aba do TAF que está tentando acessar existe e se está com o nome correto nela(Ex "1º TAF")
    3) Verificar se no lugar do indicador de ordinal 'º' foi colocado o indicador de grau '°' isso impede a leitura.''')
            input('Aperte Enter para continuar')
            cont += 1
            continue


##### iniciando a interface do programa
# para teste interno


print('-='*50)
print('''Este programa importa os lançamentos de um arquivo '.xlsx' do Excel para verificar a menção
 em relação aos índices lançados no arquivo. Para isso, o arquivo deverá estar com o mesmo nome, 
 formato(.xlsx) e colunas do modelo.''')
print('      VERSÃO EXPERIMENTAL - Sugestões ►►► TC Martins - carlos.2cmf@gmail.com')
print('-='*50)


def menu_opcoes(aba):
    while True:
        print('=-'*30)
        print(f'''OPÇÕES
        PRINCIPAIS:
            1 - Verificar erro de lançamento de menção na planilha do Excel.
            2 - Informar o nome dos militares que estão com dados faltando ou errados.
        OUTRAS OPÇÕES
            Analisar os parâmetros lançados na planilha e mostrar:
                3 - as menções, por atividade, de um militar específico.
                4 - a menção final de um militar específico.
            5 - Consultar as menções previstas a Port EME/C Ex Nº 850, DE 31 DE AGOSTO DE 2022.''')
        print(f'            6- Sair da consulta do {aba.title}?')
        try:
            op = int(input('Informe o número da opção desejada: '))
            if op == 3:
                posto = str(input('''lista dos postos: CEL, TEN CEL, CAP, 1º TEN, 2º TEN, ASP, S TEN
1º SGT, 2º SGT, 2º SGT QE, 3º SGT, CB, SD, SD EV."
Digite o Posto/Grad conforme a lista acima (Cuidado para não digitar grau '°' no lugar do indicador ordinal 'º'):

Digite o posto/graduação do militar---> ''')).upper().strip()
                nome = str(input('\nDigite o nome do militar ---> ')).upper().strip()
                print('\n')
                pg_nome = f'{posto} {nome}'
                dicio_militares = lista_mencoes(aba)
                for k,v in dicio_militares.items():
                    if k == pg_nome:
                        print(f'{k} ---> CORRIDA -> {v[0]}, FLEXÃO -> {v[1]}, ABDOMINAL -> {v[2]}, BARRA -> {v[3]}')
                print('\n')
                e = str(input('Voltar ao menu [S/N]? ')).upper()
                if e == 'S':
                    continue
                else:
                    return
            elif op == 4:
                posto = str(input('''lista dos postos: CEL, TEN CEL, CAP, 1º TEN, 2º TEN, ASP, S TEN
    1º SGT, 2º SGT, 2º SGT QE, 3º SGT, CB, SD, SD EV."
    Digite o Posto/Grad conforme a lista acima (Cuidado para não digitar grau '°' no lugar do indicador ordinal 'º'):

Digite o posto/graduação do militar---> ''')).upper().strip()
                nome = str(input('\nDigite o nome do militar ---> ')).upper().strip()
                print('\n')
                pg_nome = f'{posto} {nome}'
                mencoesfinais = mencao_final(lista_mencoes(aba))
                for k,v in mencoesfinais.items():
                    if k == f'{posto} {nome}':
                        print(f'Pelos índices lançados, a menção do(a) {k} é →   {v}')
                        print('\n')
                e = str(input('Voltar ao menu [S/N]? ')).upper().strip()
                if e == 'S':
                    continue
                else:
                    return
                    
            elif op == 2:
                erros = verifica_erros_lancamento(lista_mencoes(aba))
                if erros == {}:
                    print('*'*20)
                    print('NENHUM DADO FALTANDO!')
                    print('*'*20)
                else:
                    for k,v in erros.items():
                        print(f'{k} -> {v}')
                    print('\n')
                e = str(input('Voltar ao menu [S/N]? ')).upper().strip()
                if e == 'S':
                    continue
                else:
                    return
            
            elif op == 1:
                erros_lancamento(mencao_lancada(aba),mencao_final_limpa(limpa_mencao(lista_mencoes(aba),verifica_erros_lancamento(lista_mencoes(aba)))))
                e = str(input('Voltar ao menu [S/N]? ')).upper().strip()
                if e == 'S':
                    continue
                else:
                    return
            elif op == 5:
                while True: 
                    lem = str(input('''Informe a Linha de Ensino Militar:
                    Digite B - para bélica
                    Digite CT - para as demais
                    --> ''')).upper().strip()
                    if lem not in ['B','CT']:
                        print('''Opção inválida - Digite 'B' para bélica e 'CT' - para as demais''')
                        continue
                    else:
                        break
                while True:
                    segmento = str(input(''''Informe o segmento:
                    Digite M - para Masculino
                    Digite F - para Feminino
                    --> ''')).upper().strip()
                    if segmento not in ['M','F']:
                        print('''Opção inválida. Digite 'M' para masculino ou 'F' para feminino ''')
                        continue
                    else:
                        break
                while True:
                    atividade = str(input('''Escolha qual atividade você quer:
                    Digite 1 - para CORRIDA
                    Digite 2 - para FLEXÃO
                    Digite 3 - para ABDOMINAL
                    Digite 4 - para BARRA
                    --> ''')).strip()
                    if atividade not in ['1','2','3','4']:
                        print('Opção inválida. Digite o número correspondente à atividade.')
                        continue
                    else:
                        if atividade == '1':
                            if segmento == 'M':
                                if lem == 'B':
                                    for k,v in dicio_corrida_m_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_corrida_m_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                            else:
                                if lem == 'B':
                                    for k,v in dicio_corrida_f_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_corrida_f_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                        elif atividade == '2':
                            if segmento == 'M':
                                if lem == 'B':
                                    for k,v in dicio_flexao_m_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_flexao_m_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                            else:
                                if lem == 'B':
                                    for k,v in dicio_flexao_f_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_flexao_f_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                        elif atividade == '3':
                            if segmento == 'M':
                                if lem == 'B':
                                    for k,v in dicio_abdominal_m_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_abdominal_m_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                            else:
                                if lem == 'B':
                                    for k,v in dicio_abdominal_f_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    for k,v in dicio_abdominal_f_c.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                        else:
                            if segmento == 'M':
                                if lem == 'B':
                                    for k,v in dicio_barra_m_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    print('Não faz barra!')   
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                            else:
                                if lem == 'B':
                                    for k,v in dicio_barra_f_b.items():
                                        print(f'\nFaixa da idade -> {k}')
                                        if isinstance(v,dict):
                                            for k1,v1 in v.items():
                                                print(f'      {k1} --> {v1}')
                                        else:
                                            print('                 >=',v)    
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
                                else:
                                    print('Não faz barra!')  
                                    input('''\nAperte 'Enter'para continuar''')   
                                    break
            elif op == 6:
                return
            else:
                print('Número inválido.')
                input('\nAperte Enter para continuar ')
        except:
            print('Opção inválida.')
            input('\nAperte Enter para continuar ')

def main():
    diretorio_atual = Path.cwd() #pega o diretório atual
    nome_arquivo = diretorio_atual/'PLANILHA TAF.xlsx'#pega o nome do arquivo
    menu_opcoes(pega_aba(nome_arquivo))

main()

while True:
    print('\n','*'*35)
    print('\n GOSTARIA DE CONSULTAR OUTRO TAF?\n')
    print('*'*35)
    try:
        opcao = str(input('\nSim (S) ou Não (N)? \n--> ')).strip().upper()
        if opcao in ['SIM','S']:
            main()
        elif opcao in ['NÃO','N']:
            tamanho = len('=    Bendito seja Deus, para sempre!!    =')
            print('='*tamanho)
            print('=',' '*(tamanho-4),'=')
            print('=    Bendito seja Deus, para sempre!!    =')
            print('=',' '*(tamanho-4),'=')
            print('='*tamanho)
            input('Enter')
            break
        else:
            print('\n\nOpção inválida, digite "S" para sim ou "N" para não')
    except:
        print('\n\nOpção inválida, digite "S" para sim ou "N" para não')
        continue

















    

    

